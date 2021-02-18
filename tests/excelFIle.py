import openpyxl
book=openpyxl.load_workbook("C:\\Users\\Laxman Chavan\\Desktop\\ExcelDemo.xlsx")
sheet=book.active
Dict={}
cell=sheet.cell(row=2, column=3)
print(cell.value)

# To write something into xl sheet

sheet.cell(row=1,column=4).value="Gender"
#print(sheet.cell(row=1,column=4).value)

RowCount=sheet.max_row
ColumnCount=sheet.max_column
print(RowCount)
print(ColumnCount)

# to print all element in excle sheet for that perpose we use for loop concept here

for i in range(1,RowCount+1): #this for loop for row
    for j in range(1,ColumnCount+1): # This for loop is for column
        print(sheet.cell(row=i,column=j).value)

#problem suppose w want to retrive only one row ie test2 so how can we fetch that value only lets see
print("Only sortd object is ============")
for i in range(1,RowCount+1):
    if sheet.cell(row=i,column=1).value =="testcase2": #the problm is that here we are printing test case name aslo which is not requried
        for j in range(2,ColumnCount+1):   #To skip the test case name and get actual data start j loop with 2
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)